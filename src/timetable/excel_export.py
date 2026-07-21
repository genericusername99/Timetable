"""
Entry and exit point of all excel communication. Read in excel entry and export to excel.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from timetable.constants import AFTERNOON
from timetable.model import SchoolClass, Teacher
from timetable.enums.teacher_role import TeacherRole

WEEKDAY_NAMES = {
    "Mo": "Montag",
    "Di": "Dienstag",
    "Mi": "Mittwoch",
    "Do": "Donnerstag",
    "Fr": "Freitag",
}
WEEKDAY_CODES = {name: code for code, name in WEEKDAY_NAMES.items()}

# Colors extracted from data/format.xlsx's teacher legend (two of them -- Knoell
# and Spaeth -- used theme colors there; resolved here to their concrete RGB via
# the workbook's theme + tint, so this palette reproduces the same look).
TEACHER_COLOR_PALETTE = [
    "FF0000",  # red
    "FFFF00",  # yellow
    "00B050",  # green
    "7030A0",  # purple
    "00B0F0",  # light blue
    "FFC000",  # orange
    "92D050",  # light green
    "B4C7E7",  # pale blue
    "806000",  # olive
]

#
def load_teachers_from_excel(path: str) -> list[Teacher]:
    xls = pd.ExcelFile(path)

    df = pd.read_excel(path, sheet_name=xls.sheet_names[0]) # takes dist sheet on default, can be changed into: sheet_name="teachers" o.s.

    teachers: list[Teacher] = []

    # print("rows found:", df.columns.tolist())

    for _, row in df.iterrows():
        teacher = Teacher(
            id=str(row["id"]),
            name=str(row["name"]),

            subjects={s.strip() for s in row["subjects"].split(",")},
            max_weekly_hours=int(row["max_weekly_hours"]),
            availability={s.strip() for s in row["availability"].split(",")},

            role=TeacherRole(row["role"]),

            is_homeroom_teacher=bool(row["is_homeroom_teacher"]),
            homeroom_class=(
                str(row["homeroom_class"])
                if pd.notna(row["homeroom_class"])
                else None
            ),
        )

        teachers.append(teacher)

    return teachers


def load_classes_from_excel(
    path: str,
    sheet_name: str = "classes",
    required_hours_sheet: str = "required_hours",
) -> list[SchoolClass]:
    """
    Reads the class list (just IDs, sheet "classes") and the weekly hours each
    grade requires per subject (sheet "required_hours": one row per
    grade/subject/hours, e.g. so it's easy to add or remove a subject later
    without touching any packed string). Parallel classes in the same grade
    (e.g. "1a"/"1b") share the same requirements, looked up by grade.
    """
    classes_df = pd.read_excel(path, sheet_name=sheet_name)
    hours_df = pd.read_excel(path, sheet_name=required_hours_sheet)

    required_by_grade: dict[str, dict[str, int]] = {}
    for _, row in hours_df.iterrows():
        grade = str(row["grade"])
        subject = str(row["subject"]).strip()
        required_by_grade.setdefault(grade, {})[subject] = int(row["hours"])

    classes: list[SchoolClass] = []
    for _, row in classes_df.iterrows():
        class_id = str(row["class_id"])
        grade = class_id[:-1]

        classes.append(
            SchoolClass(
                id=class_id,
                name=class_id,
                number_of_students=0,
                required_subjects=dict(required_by_grade.get(grade, {})),
            )
        )

    return classes


def load_old_schedule_from_excel(path: str, sheet_name: str = "Stundenplan") -> dict[str, set[str]]:
    """
    Reads a previously exported schedule (same layout as
    export_colored_class_schedule_to_excel) and returns, per class, the set of
    slots that had a lesson -- used to keep a new plan's used/free slot
    pattern close to an existing, already-distributed plan instead of
    reshuffling every family's daily schedule from scratch.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    day_by_column: dict[int, str] = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == 1 and merged_range.max_row == 1:
            weekday_name = ws.cell(row=1, column=merged_range.min_col).value
            day_code = WEEKDAY_CODES.get(weekday_name)
            if day_code:
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    day_by_column[col] = day_code

    slot_by_column: dict[int, str] = {}
    for col, day in day_by_column.items():
        period = ws.cell(row=2, column=col).value
        if period is not None:
            slot_by_column[col] = f"{day}{period}"

    old_schedule: dict[str, set[str]] = {}
    for row in range(3, ws.max_row + 1):
        class_id = ws.cell(row=row, column=2).value
        if not class_id:
            continue
        old_schedule[str(class_id)] = {
            slot
            for col, slot in slot_by_column.items()
            if ws.cell(row=row, column=col).value
        }

    return old_schedule


def export_schedule_to_excel(
    result: dict[str, list[str]],
    time_slots: list[str],
    output_path: str,
    unresolved_slots: list[str] | None = None,
) -> None:
    teacher_ids = sorted(result.keys())

    grid = pd.DataFrame("", index=time_slots, columns=teacher_ids)
    for teacher_id, slots in result.items():
        for slot in slots:
            grid.loc[slot, teacher_id] = "X"

    if unresolved_slots:
        grid["Status"] = ""
        for slot in unresolved_slots:
            grid.loc[slot, "Status"] = "UNSTAFFED"

    grid.index.name = "Slot"
    grid.to_excel(output_path, sheet_name="Timetable")


def export_class_schedule_to_excel(
    schedule: dict[str, dict[str, tuple[str, str]]],
    time_slots: list[str],
    output_path: str,
    subject_shortfall: dict[tuple[str, str], int] | None = None,
) -> None:
    class_ids = sorted(schedule.keys())

    grid = pd.DataFrame("", index=time_slots, columns=class_ids)
    for class_id, lessons in schedule.items():
        for slot, (subject, _teacher_id) in lessons.items():
            grid.loc[slot, class_id] = subject

    grid.index.name = "Slot"
    grid.to_excel(output_path, sheet_name="Timetable")

    if subject_shortfall:
        shortfall_rows = [
            {"class_id": class_id, "subject": subject, "missing_hours": hours}
            for (class_id, subject), hours in subject_shortfall.items()
        ]
        shortfall_df = pd.DataFrame(shortfall_rows)
        with pd.ExcelWriter(output_path, mode="a", engine="openpyxl") as writer:
            shortfall_df.to_excel(writer, sheet_name="Shortfall", index=False)


def _assign_teacher_colors(teachers: list[Teacher]) -> dict[str, str]:
    homeroom_teachers = sorted(
        (t for t in teachers if t.is_homeroom_teacher and t.homeroom_class is not None),
        key=lambda t: t.homeroom_class,
    )
    other_teachers = sorted(
        (t for t in teachers if not (t.is_homeroom_teacher and t.homeroom_class is not None)),
        key=lambda t: t.id,
    )

    colors = {}
    for i, teacher in enumerate(homeroom_teachers + other_teachers):
        colors[teacher.id] = TEACHER_COLOR_PALETTE[i % len(TEACHER_COLOR_PALETTE)]
    return colors


def export_colored_class_schedule_to_excel(
    schedule: dict[str, dict[str, tuple[str, str]]],
    teachers: list[Teacher],
    classes: list[SchoolClass],
    time_slots: list[str],
    output_path: str,
    subject_shortfall: dict[tuple[str, str], int] | None = None,
) -> None:
    teacher_color = _assign_teacher_colors(teachers)
    teacher_by_id = {t.id: t for t in teachers}
    homeroom_teacher_id = {
        t.homeroom_class: t.id
        for t in teachers
        if t.is_homeroom_teacher and t.homeroom_class is not None
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Stundenplan"

    # Weekday headers (row 1, merged across each weekday's columns) and period
    # sub-headers (row 2), mirroring data/format.xlsx's layout. Derived from the
    # given time_slots (not re-derived from constants), so a caller-provided
    # slot list is honoured rather than silently replaced by the full week.
    # Grouped explicitly by day (rather than assuming day/period adjacency in
    # the input order) since constants.TIME_SLOTS lists all core periods first
    # and appends every day's "NU" afternoon slot only at the very end.
    day_order = []
    day_slots: dict[str, list[tuple[str, str]]] = {}
    for slot in time_slots:
        day, period = slot[:2], slot[2:]
        if day not in day_slots:
            day_order.append(day)
            day_slots[day] = []
        day_slots[day].append((slot, period))

    slot_column = {}
    col = 4  # column D
    for day in day_order:
        start_col = col
        slots_sorted = sorted(
            day_slots[day],
            key=lambda sp: (sp[1] == AFTERNOON, int(sp[1]) if sp[1] != AFTERNOON else 0),
        )
        for slot, period in slots_sorted:
            slot_column[slot] = col
            ws.cell(row=2, column=col, value=period if period == AFTERNOON else int(period))
            col += 1
        ws.cell(row=1, column=start_col, value=WEEKDAY_NAMES.get(day, day))
        if col - 1 > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)

    ws.cell(row=1, column=1, value="Klasse")

    # Class rows, grouped by grade (leading part of the class id) with a blank
    # separator row between grades, same as data/format.xlsx.
    classes_by_grade: dict[str, list[SchoolClass]] = {}
    for school_class in classes:
        classes_by_grade.setdefault(school_class.id[:-1], []).append(school_class)

    row = 3
    for grade in sorted(classes_by_grade):
        for school_class in sorted(classes_by_grade[grade], key=lambda c: c.id):
            ws.cell(row=row, column=2, value=school_class.id)

            homeroom_id = homeroom_teacher_id.get(school_class.id)
            if homeroom_id:
                cell = ws.cell(row=row, column=3, value=teacher_by_id[homeroom_id].name)
                color = teacher_color.get(homeroom_id)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

            for slot, (subject, teacher_id) in schedule.get(school_class.id, {}).items():
                if slot not in slot_column:
                    continue
                cell = ws.cell(row=row, column=slot_column[slot], value=subject)
                color = teacher_color.get(teacher_id)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

            row += 1
        row += 1  # blank separator row between grades

    # Legend for teachers without their own class row (e.g. subject specialists).
    row += 1
    ws.cell(row=row, column=1, value="Weitere Lehrkräfte")
    row += 2
    for teacher in sorted(
        (t for t in teachers if t.id not in homeroom_teacher_id.values()), key=lambda t: t.id
    ):
        cell = ws.cell(row=row, column=3, value=teacher.name)
        color = teacher_color.get(teacher.id)
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
        row += 1

    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16

    wb.save(output_path)

    if subject_shortfall:
        shortfall_rows = [
            {"class_id": class_id, "subject": subject, "missing_hours": hours}
            for (class_id, subject), hours in subject_shortfall.items()
        ]
        shortfall_df = pd.DataFrame(shortfall_rows)
        with pd.ExcelWriter(output_path, mode="a", engine="openpyxl") as writer:
            shortfall_df.to_excel(writer, sheet_name="Shortfall", index=False)


if __name__ == "__main__":
    teachers = load_teachers_from_excel("../data/example_entry.xlsx")

    for t in teachers:
        print(t)
