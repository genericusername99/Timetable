"""
Test excel_export.py: reading teachers from Excel and writing the solved schedule back out.
"""
from pathlib import Path

import openpyxl
import pandas as pd

from timetable.enums.teacher_role import TeacherRole
from timetable.excel_export import (
    export_class_schedule_to_excel,
    export_colored_class_schedule_to_excel,
    export_schedule_to_excel,
    load_classes_from_excel,
    load_old_schedule_from_excel,
    load_teachers_from_excel,
)
from timetable.model import SchoolClass, Teacher

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXAMPLE_ENTRY_PATH = PROJECT_ROOT / "data" / "example_entry.xlsx"


def test_load_teachers_from_excel_reads_real_file():
    teachers = load_teachers_from_excel(str(EXAMPLE_ENTRY_PATH))

    assert len(teachers) == 9

    homeroom_teachers = {t.homeroom_class: t for t in teachers if t.is_homeroom_teacher}
    assert set(homeroom_teachers) == {"1a", "1b", "2a", "2b", "3a", "3b", "4a", "4b"}

    class_1a_teacher = homeroom_teachers["1a"]
    assert class_1a_teacher.id == "S-R"
    assert class_1a_teacher.subjects == {
        "D", "M/LB", "SU", "M", "Sp", "E", "TW", "BK", "Mus", "D/L",
    }
    assert class_1a_teacher.max_weekly_hours == 30
    assert len(class_1a_teacher.availability) == 35

    specialist = next(t for t in teachers if t.id == "Späth")
    assert specialist.subjects == {"Rel"}
    assert specialist.max_weekly_hours == 31
    assert specialist.is_homeroom_teacher is False
    assert specialist.homeroom_class is None


def test_load_classes_from_excel_reads_real_file():
    classes = load_classes_from_excel(str(EXAMPLE_ENTRY_PATH))

    assert len(classes) == 8
    class_ids = {c.id for c in classes}
    assert class_ids == {"1a", "1b", "2a", "2b", "3a", "3b", "4a", "4b"}

    class_1a = next(c for c in classes if c.id == "1a")
    assert class_1a.required_subjects == {
        "D": 6, "M/LB": 2, "SU": 3, "M": 4, "BK": 1, "Mus": 1, "Sp": 2, "D/L": 1,
    }

    class_2a = next(c for c in classes if c.id == "2a")
    assert class_2a.required_subjects["Rel"] == 2
    assert class_2a.required_subjects["TW"] == 1

    # Parallel classes share the same grade-level requirements (looked up from
    # the "required_hours" table, not duplicated per class).
    class_1b = next(c for c in classes if c.id == "1b")
    assert class_1b.required_subjects == class_1a.required_subjects


def test_export_schedule_to_excel_creates_grid(tmp_path):
    result = {
        "frabai": ["Mo1", "Mo3"],
        "balu": ["Mo2"],
    }
    time_slots = ["Mo1", "Mo2", "Mo3", "Mo4"]
    output_path = tmp_path / "schedule.xlsx"

    export_schedule_to_excel(result, time_slots, str(output_path))

    df = pd.read_excel(output_path, sheet_name="Timetable", index_col=0)

    assert list(df.index) == time_slots
    assert set(df.columns) == {"frabai", "balu"}

    assert df.loc["Mo1", "frabai"] == "X"
    assert df.loc["Mo3", "frabai"] == "X"
    assert df.loc["Mo2", "balu"] == "X"
    assert pd.isna(df.loc["Mo4", "frabai"]) or df.loc["Mo4", "frabai"] == ""
    assert pd.isna(df.loc["Mo1", "balu"]) or df.loc["Mo1", "balu"] == ""


def test_export_schedule_to_excel_marks_unresolved_slots(tmp_path):
    result = {"frabai": ["Mo1"]}
    time_slots = ["Mo1", "Mo2", "Mo3"]
    output_path = tmp_path / "schedule.xlsx"

    export_schedule_to_excel(
        result, time_slots, str(output_path), unresolved_slots=["Mo2", "Mo3"]
    )

    df = pd.read_excel(output_path, sheet_name="Timetable", index_col=0)

    assert "Status" in df.columns
    assert df.loc["Mo2", "Status"] == "UNSTAFFED"
    assert df.loc["Mo3", "Status"] == "UNSTAFFED"
    assert pd.isna(df.loc["Mo1", "Status"]) or df.loc["Mo1", "Status"] == ""


def test_export_class_schedule_to_excel_creates_grid(tmp_path):
    schedule = {"1a": {"Mo1": ("D", "S-R"), "Mo2": ("M", "S-R")}}
    time_slots = ["Mo1", "Mo2", "Mo3"]
    output_path = tmp_path / "schedule.xlsx"

    export_class_schedule_to_excel(schedule, time_slots, str(output_path))

    df = pd.read_excel(output_path, sheet_name="Timetable", index_col=0)
    assert df.loc["Mo1", "1a"] == "D"
    assert df.loc["Mo2", "1a"] == "M"
    assert pd.isna(df.loc["Mo3", "1a"]) or df.loc["Mo3", "1a"] == ""


def _make_teacher(id, subjects, is_homeroom_teacher=False, homeroom_class=None):
    return Teacher(
        id=id,
        name=id,
        subjects=set(subjects),
        max_weekly_hours=30,
        availability={"Mo1", "MoNU"},
        role=TeacherRole.FULL,
        is_homeroom_teacher=is_homeroom_teacher,
        homeroom_class=homeroom_class,
    )


def test_export_colored_class_schedule_to_excel_matches_format_style(tmp_path):
    homeroom = _make_teacher("S-R", {"D"}, is_homeroom_teacher=True, homeroom_class="1a")
    specialist = _make_teacher("Späth", {"Rel"})
    teachers = [homeroom, specialist]
    classes = [SchoolClass(id="1a", name="1a", number_of_students=20, required_subjects={"D": 1, "Rel": 1})]
    schedule = {"1a": {"Mo1": ("D", "S-R"), "MoNU": ("Rel", "Späth")}}
    time_slots = ["Mo1", "MoNU"]
    output_path = tmp_path / "schedule.xlsx"

    export_colored_class_schedule_to_excel(schedule, teachers, classes, time_slots, str(output_path))

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Stundenplan"]

    assert ws["B3"].value == "1a"
    assert ws["C3"].value == "S-R"
    homeroom_color = ws["C3"].fill.fgColor.rgb
    assert ws["D3"].value == "D"
    assert ws["D3"].fill.fgColor.rgb == homeroom_color

    # The specialist teaches "Rel" for this class -- her cell must carry HER
    # color, distinct from the homeroom teacher's, mirroring format.xlsx where
    # Knoell's cells in another class's row are colored in Knoell's color.
    assert ws["E3"].value == "Rel"
    specialist_color = ws["E3"].fill.fgColor.rgb
    assert specialist_color != homeroom_color

    # Späth has no class of her own, so she shows up in the legend section only.
    assert "Späth" in [row[0].value for row in ws.iter_rows(min_col=3, max_col=3) if row[0].value]

    # Required-hours-per-grade is now only in the input file, not re-printed
    # into the output.
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert not any(isinstance(v, str) and v.startswith("Klasse ") for v in all_values)


def test_load_old_schedule_from_excel_round_trips_with_export(tmp_path):
    homeroom = _make_teacher("S-R", {"D"}, is_homeroom_teacher=True, homeroom_class="1a")
    teachers = [homeroom]
    classes = [SchoolClass(id="1a", name="1a", number_of_students=20, required_subjects={"D": 2})]
    schedule = {"1a": {"Mo1": ("D", "S-R"), "MoNU": ("D", "S-R")}}
    time_slots = ["Mo1", "Mo2", "MoNU"]
    output_path = tmp_path / "old_plan.xlsx"

    export_colored_class_schedule_to_excel(schedule, teachers, classes, time_slots, str(output_path))

    old_schedule = load_old_schedule_from_excel(str(output_path))

    assert old_schedule["1a"] == {"Mo1", "MoNU"}
